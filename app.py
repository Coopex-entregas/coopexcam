
import os
import secrets
import json
from datetime import datetime, timezone
from collections import defaultdict
from io import BytesIO

import qrcode
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, join_room, emit, disconnect
from sqlalchemy import func, inspect, text
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app = Flask(__name__, instance_relative_config=True)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'coopexcam-secret')
db_url = os.getenv('DATABASE_URL')
if db_url and db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///' + os.path.join(BASE_DIR, 'instance', 'coopexcam.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
socketio = SocketIO(app, cors_allowed_origins='*', async_mode='threading', manage_session=False)

ADMIN_LOGIN = 'coopex'
ADMIN_PASSWORD = '05289'


class MeetingRoom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(120), nullable=False)
    code = db.Column(db.String(40), nullable=False, unique=True, index=True)
    invite_token = db.Column(db.String(80), nullable=False, unique=True, index=True)
    status = db.Column(db.String(20), default='open', nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)
    ended_at = db.Column(db.DateTime)
    allow_microphone = db.Column(db.Boolean, default=True, nullable=False)
    allow_camera = db.Column(db.Boolean, default=True, nullable=False)
    speech_mode = db.Column(db.String(20), default='controlled', nullable=False)
    summary_text = db.Column(db.Text, default='', nullable=False)
    decisions_text = db.Column(db.Text, default='', nullable=False)


class Participant(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    room_id = db.Column(db.Integer, db.ForeignKey('meeting_room.id'), nullable=False, index=True)
    full_name = db.Column(db.String(160), nullable=False)
    display_name = db.Column(db.String(80), nullable=False)
    join_token = db.Column(db.String(80), nullable=False, unique=True, index=True)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)
    is_eligible = db.Column(db.Boolean, default=False, nullable=False)
    can_speak = db.Column(db.Boolean, default=False, nullable=False)
    mic_blocked = db.Column(db.Boolean, default=False, nullable=False)
    cam_blocked = db.Column(db.Boolean, default=False, nullable=False)
    online = db.Column(db.Boolean, default=False, nullable=False)
    removed = db.Column(db.Boolean, default=False, nullable=False)
    joined_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)
    left_at = db.Column(db.DateTime)


class AttendanceLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    room_id = db.Column(db.Integer, db.ForeignKey('meeting_room.id'), nullable=False, index=True)
    participant_id = db.Column(db.Integer, db.ForeignKey('participant.id'), nullable=False, index=True)
    entered_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)
    exited_at = db.Column(db.DateTime)


class VoteSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    room_id = db.Column(db.Integer, db.ForeignKey('meeting_room.id'), nullable=False, index=True)
    title = db.Column(db.String(200), nullable=False)
    options_csv = db.Column(db.Text, nullable=False)
    rule = db.Column(db.String(40), default='simple_majority', nullable=False)
    secret = db.Column(db.Boolean, default=False, nullable=False)
    active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)
    ended_at = db.Column(db.DateTime)

    @property
    def options(self):
        raw = self.options_csv or ''
        return [x.strip() for x in raw.split('|') if x.strip()]


class VoteRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vote_session_id = db.Column(db.Integer, db.ForeignKey('vote_session.id'), nullable=False, index=True)
    participant_id = db.Column(db.Integer, db.ForeignKey('participant.id'), nullable=False, index=True)
    option = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), nullable=False)
    __table_args__ = (db.UniqueConstraint('vote_session_id', 'participant_id', name='uq_vote_once'),)


room_runtime = defaultdict(lambda: {
    'hands': [],
    'speaker_id': None,
    'selected_id': None,
    'screen_share_id': None,
})

sid_to_participant = {}


def utcnow():
    return datetime.now(timezone.utc)


def room_channel(code: str) -> str:
    return f'room:{code}'


def participant_channel(p: Participant) -> str:
    return f'participant:{p.join_token}'


def codeify(value: str) -> str:
    cleaned = ''.join(ch for ch in value.upper() if ch.isalnum())
    return cleaned[:16] or secrets.token_hex(3).upper()


def unique_code(base: str) -> str:
    code = codeify(base)
    candidate = code
    i = 2
    while MeetingRoom.query.filter_by(code=candidate).first():
        candidate = f'{code[:12]}{i}'
        i += 1
    return candidate


def _column_names(table_name: str):
    inspector = inspect(db.engine)
    if table_name not in inspector.get_table_names():
        return set()
    return {c['name'] for c in inspector.get_columns(table_name)}


def ensure_schema():
    inspector = inspect(db.engine)

    if 'meeting_room' in inspector.get_table_names():
        cols = _column_names('meeting_room')
        statements = []
        wanted = {
            'invite_token': "ALTER TABLE meeting_room ADD COLUMN invite_token VARCHAR(80)",
            'status': "ALTER TABLE meeting_room ADD COLUMN status VARCHAR(20) DEFAULT 'open'",
            'ended_at': "ALTER TABLE meeting_room ADD COLUMN ended_at TIMESTAMP NULL",
            'allow_microphone': "ALTER TABLE meeting_room ADD COLUMN allow_microphone BOOLEAN DEFAULT TRUE",
            'allow_camera': "ALTER TABLE meeting_room ADD COLUMN allow_camera BOOLEAN DEFAULT TRUE",
            'speech_mode': "ALTER TABLE meeting_room ADD COLUMN speech_mode VARCHAR(20) DEFAULT 'controlled'",
            'summary_text': "ALTER TABLE meeting_room ADD COLUMN summary_text TEXT DEFAULT ''",
            'decisions_text': "ALTER TABLE meeting_room ADD COLUMN decisions_text TEXT DEFAULT ''",
        }
        for key, stmt in wanted.items():
            if key not in cols:
                statements.append(stmt)
        for stmt in statements:
            db.session.execute(text(stmt))
        if statements:
            db.session.commit()

        cols = _column_names('meeting_room')
        if 'admin_user_id' in cols:
            try:
                db.session.execute(text("ALTER TABLE meeting_room ALTER COLUMN admin_user_id DROP NOT NULL"))
                db.session.commit()
            except Exception:
                db.session.rollback()

        db.session.execute(text("UPDATE meeting_room SET summary_text = '' WHERE summary_text IS NULL"))
        db.session.execute(text("UPDATE meeting_room SET decisions_text = '' WHERE decisions_text IS NULL"))
        if 'invite_token' in cols or 'invite_token' in wanted:
            db.session.execute(text("UPDATE meeting_room SET invite_token = md5(random()::text || clock_timestamp()::text) WHERE invite_token IS NULL OR invite_token = ''"))
        db.session.commit()

    if 'participant' in inspector.get_table_names():
        cols = _column_names('participant')
        statements = []
        wanted = {
            'is_eligible': "ALTER TABLE participant ADD COLUMN is_eligible BOOLEAN DEFAULT FALSE",
            'can_speak': "ALTER TABLE participant ADD COLUMN can_speak BOOLEAN DEFAULT FALSE",
            'mic_blocked': "ALTER TABLE participant ADD COLUMN mic_blocked BOOLEAN DEFAULT FALSE",
            'cam_blocked': "ALTER TABLE participant ADD COLUMN cam_blocked BOOLEAN DEFAULT FALSE",
            'online': "ALTER TABLE participant ADD COLUMN online BOOLEAN DEFAULT FALSE",
            'left_at': "ALTER TABLE participant ADD COLUMN left_at TIMESTAMP NULL",
            'removed': "ALTER TABLE participant ADD COLUMN removed BOOLEAN DEFAULT FALSE",
        }
        for key, stmt in wanted.items():
            if key not in cols:
                statements.append(stmt)
        for stmt in statements:
            db.session.execute(text(stmt))
        if statements:
            db.session.commit()
        db.session.execute(text("UPDATE participant SET removed = FALSE WHERE removed IS NULL"))
        db.session.commit()

    if 'vote_session' in inspector.get_table_names():
        cols = _column_names('vote_session')
        statements = []
        wanted = {
            'options_csv': "ALTER TABLE vote_session ADD COLUMN options_csv TEXT",
            'rule': "ALTER TABLE vote_session ADD COLUMN rule VARCHAR(40) DEFAULT 'simple_majority'",
            'secret': "ALTER TABLE vote_session ADD COLUMN secret BOOLEAN DEFAULT FALSE",
            'active': "ALTER TABLE vote_session ADD COLUMN active BOOLEAN DEFAULT TRUE",
            'ended_at': "ALTER TABLE vote_session ADD COLUMN ended_at TIMESTAMP NULL",
        }
        for key, stmt in wanted.items():
            if key not in cols:
                statements.append(stmt)
        for stmt in statements:
            db.session.execute(text(stmt))
        if statements:
            db.session.commit()
        cols = _column_names('vote_session')
        if 'options_json' in cols:
            rows = db.session.execute(text("""
                SELECT id, options_json FROM vote_session
                WHERE (options_csv IS NULL OR options_csv = '') AND options_json IS NOT NULL
            """)).fetchall()
            for row in rows:
                raw = row[1]
                try:
                    parsed = json.loads(raw) if isinstance(raw, str) else raw
                except Exception:
                    parsed = None
                if isinstance(parsed, list):
                    csv_value = '|'.join(str(x).strip() for x in parsed if str(x).strip())
                else:
                    csv_value = 'Sim|Não|Abstenção'
                db.session.execute(text("UPDATE vote_session SET options_csv = :csv WHERE id = :id"), {'csv': csv_value, 'id': row[0]})
            db.session.commit()
        db.session.execute(text("UPDATE vote_session SET options_csv = 'Sim|Não|Abstenção' WHERE options_csv IS NULL OR options_csv = ''"))
        db.session.commit()


def active_participants_query(room_id):
    return Participant.query.filter_by(room_id=room_id, removed=False)


def tally_vote(vote: VoteSession, room: MeetingRoom):
    eligible_count = active_participants_query(room.id).filter_by(is_eligible=True).count()
    online_count = active_participants_query(room.id).filter_by(online=True).count()
    rows = VoteRecord.query.filter_by(vote_session_id=vote.id).all()
    voted = len(rows)
    counts = {opt: 0 for opt in vote.options}
    for row in rows:
        counts[row.option] = counts.get(row.option, 0) + 1
    total_for_percent = max(voted, 1)
    percentages = {opt: round((counts.get(opt, 0) / total_for_percent) * 100, 1) if voted else 0 for opt in vote.options}
    result = 'Em andamento'
    if not vote.active:
        if vote.rule == 'simple_majority':
            approved = counts.get('Sim', 0) > counts.get('Não', 0)
        elif vote.rule == 'absolute_majority':
            approved = counts.get('Sim', 0) >= (eligible_count // 2 + 1)
        else:
            approved = counts.get('Sim', 0) >= ((online_count * 2 + 2) // 3)
        result = 'Aprovada' if approved else 'Não aprovada'
    details = []
    if not vote.secret:
        for row in rows:
            p = Participant.query.get(row.participant_id)
            if p:
                details.append({'name': p.full_name, 'option': row.option})
    return {
        'presentes': online_count,
        'aptos': eligible_count,
        'votaram': voted,
        'faltam': max(eligible_count - voted, 0),
        'counts': counts,
        'percentages': percentages,
        'result': result,
        'details': details,
    }


def room_state(room: MeetingRoom):
    runtime = room_runtime[room.code]
    participants = active_participants_query(room.id).order_by(Participant.is_admin.desc(), Participant.display_name.asc()).all()
    hands = [pid for pid in runtime['hands'] if any(p.id == pid for p in participants)]
    runtime['hands'] = hands
    rows = []
    for p in participants:
        rows.append({
            'id': p.id,
            'full_name': p.full_name,
            'display_name': p.display_name,
            'is_admin': p.is_admin,
            'online': p.online,
            'is_eligible': p.is_eligible,
            'can_speak': p.can_speak or p.is_admin or room.speech_mode == 'free',
            'mic_blocked': p.mic_blocked or (not room.allow_microphone and not p.is_admin),
            'cam_blocked': p.cam_blocked or (not room.allow_camera and not p.is_admin),
            'hand_raised': p.id in hands,
            'selected': runtime['selected_id'] == p.id,
            'speaking': runtime['speaker_id'] == p.id,
        })
    active_vote = VoteSession.query.filter_by(room_id=room.id, active=True).order_by(VoteSession.id.desc()).first()
    vote_data = None
    if active_vote:
        vote_data = tally_vote(active_vote, room)
        vote_data.update({
            'id': active_vote.id,
            'title': active_vote.title,
            'options': active_vote.options,
            'secret': active_vote.secret,
            'rule': active_vote.rule,
            'active': active_vote.active,
        })
    return {
        'room': {
            'id': room.id,
            'title': room.title,
            'code': room.code,
            'invite_url': url_for('join_token_page', token=room.invite_token, _external=True),
            'camera_url': url_for('camera_companion', code=room.code, _external=True),
            'camera_qr_url': url_for('camera_qr', code=room.code),
            'status': room.status,
            'allow_microphone': room.allow_microphone,
            'allow_camera': room.allow_camera,
            'speech_mode': room.speech_mode,
            'selected_id': runtime['selected_id'],
            'speaker_id': runtime['speaker_id'],
            'screen_share_id': runtime.get('screen_share_id'),
            'summary_text': room.summary_text or '',
            'decisions_text': room.decisions_text or '',
            'created_at': room.created_at.strftime('%d/%m/%Y %H:%M') if room.created_at else '',
            'ended_at': room.ended_at.strftime('%d/%m/%Y %H:%M') if room.ended_at else '',
        },
        'participants': rows,
        'vote': vote_data,
        'hands': hands,
    }


def current_admin():
    return session.get('admin_ok') is True


def emit_room_state(code: str):
    room = MeetingRoom.query.filter_by(code=code).first()
    if room:
        socketio.emit('room_state', room_state(room), to=room_channel(code))


def get_or_create_named_participant(room, full_name, display_name, can_speak=False):
    existing = Participant.query.filter_by(room_id=room.id, full_name=full_name, display_name=display_name, removed=False).order_by(Participant.id.desc()).first()
    if existing:
        return existing
    participant = Participant(
        room_id=room.id,
        full_name=full_name,
        display_name=display_name,
        join_token=secrets.token_urlsafe(24),
        can_speak=can_speak,
    )
    db.session.add(participant)
    db.session.commit()
    return participant


@app.route('/')
def home():
    return redirect(url_for('dashboard') if current_admin() else url_for('admin_login'))


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        login = request.form.get('login', '').strip()
        password = request.form.get('password', '').strip()
        if login == ADMIN_LOGIN and password == ADMIN_PASSWORD:
            session['admin_ok'] = True
            return redirect(url_for('dashboard'))
        flash('Login inválido.')
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))


@app.route('/admin/dashboard')
def dashboard():
    if not current_admin():
        return redirect(url_for('admin_login'))
    rooms = MeetingRoom.query.order_by(MeetingRoom.created_at.desc()).all()
    return render_template('dashboard.html', rooms=rooms)


@app.route('/admin/history')
def admin_history():
    if not current_admin():
        return redirect(url_for('admin_login'))
    rooms = MeetingRoom.query.order_by(MeetingRoom.created_at.desc()).all()
    return render_template('history.html', rooms=rooms)


@app.route('/admin/create_room', methods=['POST'])
def create_room():
    if not current_admin():
        return redirect(url_for('admin_login'))
    title = request.form.get('title', '').strip() or 'Reunião CoopexCam'
    desired = request.form.get('code', '').strip()
    code = unique_code(desired or title)
    speech_mode = request.form.get('speech_mode', 'controlled') or 'controlled'
    room = MeetingRoom(
        title=title,
        code=code,
        invite_token=secrets.token_urlsafe(24),
        status='open',
        allow_microphone=True,
        allow_camera=True,
        speech_mode=speech_mode
    )
    db.session.add(room)
    db.session.flush()
    admin_p = Participant(
        room_id=room.id,
        full_name='Administrador CoopexCam',
        display_name='Administrador',
        join_token=secrets.token_urlsafe(24),
        is_admin=True,
        can_speak=True
    )
    db.session.add(admin_p)
    db.session.commit()
    return redirect(url_for('admin_room', code=room.code))


@app.route('/join/<token>', methods=['GET', 'POST'])
def join_token_page(token):
    room = MeetingRoom.query.filter_by(invite_token=token).first_or_404()
    if room.status != 'open':
        return render_template('join_closed.html', room=room)
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        display_name = request.form.get('display_name', '').strip()
        if not full_name or not display_name:
            flash('Preencha nome completo e nome de exibição.')
        else:
            blocked = Participant.query.filter_by(room_id=room.id, full_name=full_name, removed=True).first()
            if blocked:
                flash('Este participante foi removido da sala e não pode entrar novamente.')
            else:
                participant = get_or_create_named_participant(room, full_name, display_name, can_speak=(room.speech_mode == 'free'))
                return redirect(url_for('participant_room_page', join_token=participant.join_token))
    return render_template('join_form.html', room=room)


@app.route('/camera/<code>')
def camera_companion(code):
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    participant = get_or_create_named_participant(room, 'Câmera do Admin', 'Câmera do Admin', can_speak=True)
    return redirect(url_for('participant_room_page', join_token=participant.join_token))


@app.route('/camera-qr/<code>')
def camera_qr(code):
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    qr = qrcode.QRCode(box_size=8, border=2)
    qr.add_data(url_for('camera_companion', code=room.code, _external=True))
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    bio = BytesIO()
    img.save(bio, format='PNG')
    bio.seek(0)
    return send_file(bio, mimetype='image/png')


@app.route('/admin/room/<code>')
def admin_room(code):
    if not current_admin():
        return redirect(url_for('admin_login'))
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    admin_p = Participant.query.filter_by(room_id=room.id, is_admin=True).first()
    return render_template('admin_room.html', room=room, state=room_state(room), join_token=admin_p.join_token)


@app.route('/room/<join_token>')
def participant_room_page(join_token):
    participant = Participant.query.filter_by(join_token=join_token).first_or_404()
    room = MeetingRoom.query.get_or_404(participant.room_id)
    if participant.removed:
        flash('Você foi removido da sala.')
        return redirect(url_for('home'))
    return render_template('participant_room.html', room=room, state=room_state(room), join_token=join_token, participant=participant)


@app.route('/admin/api/room/<code>/toggle_status', methods=['POST'])
def toggle_status(code):
    if not current_admin():
        return jsonify(ok=False), 403
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    room.status = 'ended' if room.status == 'open' else 'open'
    room.ended_at = utcnow() if room.status == 'ended' else None
    if room.status == 'ended':
        Participant.query.filter_by(room_id=room.id).update({'online': False, 'left_at': utcnow()})
    db.session.commit()
    emit_room_state(code)
    return jsonify(ok=True, status=room.status)


@app.route('/admin/api/room/<code>/delete', methods=['POST'])
def delete_room(code):
    if not current_admin():
        return jsonify(ok=False), 403
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    participant_ids = [p.id for p in Participant.query.filter_by(room_id=room.id).all()]
    if participant_ids:
        VoteRecord.query.filter(VoteRecord.participant_id.in_(participant_ids)).delete(synchronize_session=False)
        AttendanceLog.query.filter(AttendanceLog.participant_id.in_(participant_ids)).delete(synchronize_session=False)
    VoteSession.query.filter_by(room_id=room.id).delete(synchronize_session=False)
    Participant.query.filter_by(room_id=room.id).delete(synchronize_session=False)
    db.session.delete(room)
    db.session.commit()
    room_runtime.pop(code, None)
    return jsonify(ok=True)


@app.route('/admin/api/room/<code>/participant/<int:pid>', methods=['POST'])
def participant_action(code, pid):
    if not current_admin():
        return jsonify(ok=False), 403
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    p = Participant.query.filter_by(room_id=room.id, id=pid).first_or_404()
    action = (request.json or {}).get('action')
    runtime = room_runtime[room.code]
    if action == 'toggle_eligible':
        p.is_eligible = not p.is_eligible
    elif action == 'allow_speak':
        p.can_speak = True
        if p.id in runtime['hands']:
            runtime['hands'].remove(p.id)
        runtime['selected_id'] = p.id
    elif action == 'block_mic':
        p.mic_blocked = not p.mic_blocked
    elif action == 'block_cam':
        p.cam_blocked = not p.cam_blocked
    elif action == 'remove':
        channel = participant_channel(p)
        p.online = False
        p.left_at = utcnow()
        p.removed = True
        db.session.commit()
        socketio.emit('removed', {'reason': 'Você foi removido da sala.'}, to=channel)
        emit_room_state(room.code)
        return jsonify(ok=True)
    elif action == 'spotlight':
        runtime['selected_id'] = p.id if runtime['selected_id'] != p.id else None
    db.session.commit()
    emit_room_state(room.code)
    return jsonify(ok=True)


@app.route('/admin/api/room/<code>/bulk', methods=['POST'])
def bulk_action(code):
    if not current_admin():
        return jsonify(ok=False), 403
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    action = (request.json or {}).get('action')
    participants = active_participants_query(room.id).filter(Participant.is_admin == False).all()
    if action == 'eligible_all':
        for p in participants:
            p.is_eligible = True
    elif action == 'eligible_none':
        for p in participants:
            p.is_eligible = False
    elif action == 'mute_all':
        for p in participants:
            p.mic_blocked = True
    elif action == 'camera_off_all':
        for p in participants:
            p.cam_blocked = True
    elif action == 'mic_global':
        room.allow_microphone = not room.allow_microphone
    elif action == 'cam_global':
        room.allow_camera = not room.allow_camera
    elif action == 'speech_mode':
        room.speech_mode = 'free' if room.speech_mode == 'controlled' else 'controlled'
        if room.speech_mode == 'free':
            for p in participants:
                p.can_speak = True
    db.session.commit()
    emit_room_state(room.code)
    return jsonify(ok=True)


@app.route('/admin/api/room/<code>/notes', methods=['POST'])
def save_notes(code):
    if not current_admin():
        return jsonify(ok=False), 403
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    payload = request.json or {}
    room.summary_text = (payload.get('summary_text') or '').strip()
    room.decisions_text = (payload.get('decisions_text') or '').strip()
    db.session.commit()
    emit_room_state(room.code)
    return jsonify(ok=True)


@app.route('/admin/api/room/<code>/vote', methods=['POST'])
def create_vote(code):
    if not current_admin():
        return jsonify(ok=False), 403
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    active = VoteSession.query.filter_by(room_id=room.id, active=True).first()
    if active:
        return jsonify(ok=False, message='Já existe votação em andamento.'), 400
    payload = request.json or {}
    options = [str(x).strip() for x in (payload.get('options') or ['Sim', 'Não', 'Abstenção']) if str(x).strip()]
    if not payload.get('title'):
        return jsonify(ok=False, message='Informe o título da votação.'), 400
    vote = VoteSession(
        room_id=room.id,
        title=payload.get('title', 'Votação'),
        options_csv='|'.join(options),
        rule=payload.get('rule', 'simple_majority'),
        secret=bool(payload.get('secret')),
        active=True,
    )
    db.session.add(vote)
    db.session.commit()
    emit_room_state(room.code)
    return jsonify(ok=True)


@app.route('/admin/api/room/<code>/vote/end', methods=['POST'])
def end_vote(code):
    if not current_admin():
        return jsonify(ok=False), 403
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    vote = VoteSession.query.filter_by(room_id=room.id, active=True).first()
    if not vote:
        return jsonify(ok=False, message='Sem votação ativa.'), 400
    vote.active = False
    vote.ended_at = utcnow()
    db.session.commit()
    emit_room_state(room.code)
    return jsonify(ok=True)


@app.route('/api/vote/<join_token>', methods=['POST'])
def cast_vote(join_token):
    p = Participant.query.filter_by(join_token=join_token).first_or_404()
    room = MeetingRoom.query.get_or_404(p.room_id)
    if p.removed:
        return jsonify(ok=False, message='Participante removido.'), 403
    vote = VoteSession.query.filter_by(room_id=room.id, active=True).first()
    if not vote:
        return jsonify(ok=False, message='Sem votação ativa.'), 400
    if not p.is_eligible:
        return jsonify(ok=False, message='Você não está apto a votar.'), 400
    option = (request.json or {}).get('option')
    if option not in vote.options:
        return jsonify(ok=False, message='Opção inválida.'), 400
    already = VoteRecord.query.filter_by(vote_session_id=vote.id, participant_id=p.id).first()
    if already:
        return jsonify(ok=False, message='Voto já registrado.'), 400
    db.session.add(VoteRecord(vote_session_id=vote.id, participant_id=p.id, option=option))
    db.session.commit()
    emit_room_state(room.code)
    return jsonify(ok=True, message='Voto registrado.')


@app.route('/admin/export/<code>.xlsx')
def export_xlsx(code):
    if not current_admin():
        return redirect(url_for('admin_login'))
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Presença'
    ws.append(['Nome completo', 'Primeira entrada', 'Última saída'])
    sub = db.session.query(
        AttendanceLog.participant_id,
        func.min(AttendanceLog.entered_at).label('first_in'),
        func.max(AttendanceLog.exited_at).label('last_out')
    ).filter_by(room_id=room.id).group_by(AttendanceLog.participant_id).all()
    for item in sub:
        p = Participant.query.get(item.participant_id)
        if p:
            ws.append([p.full_name, str(item.first_in or ''), str(item.last_out or '')])

    notes = wb.create_sheet('Resumo')
    notes.append(['Resumo da reunião'])
    notes.append([room.summary_text or ''])
    notes.append([])
    notes.append(['Decisões'])
    notes.append([room.decisions_text or ''])

    vote = VoteSession.query.filter_by(room_id=room.id).order_by(VoteSession.id.desc()).first()
    if vote:
        tally = tally_vote(vote, room)
        vs = wb.create_sheet('Votação')
        vs.append(['Título', vote.title])
        vs.append(['Resultado', tally['result']])
        vs.append(['Presentes', tally['presentes']])
        vs.append(['Aptos', tally['aptos']])
        vs.append(['Votaram', tally['votaram']])
        vs.append(['Faltam', tally['faltam']])
        vs.append([])
        vs.append(['Opção', 'Total', 'Percentual'])
        for opt, count in tally['counts'].items():
            vs.append([opt, count, f"{tally['percentages'].get(opt, 0)}%"])
        if tally['details']:
            vs.append([])
            vs.append(['Nome', 'Voto'])
            for detail in tally['details']:
                vs.append([detail['name'], detail['option']])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f'{room.code}.xlsx')


@app.route('/admin/export/<code>.pdf')
def export_pdf(code):
    if not current_admin():
        return redirect(url_for('admin_login'))
    room = MeetingRoom.query.filter_by(code=code).first_or_404()
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    y = 800
    c.setFont('Helvetica-Bold', 14)
    c.drawString(40, y, f'CoopexCam - {room.title}')
    y -= 24
    c.setFont('Helvetica', 10)
    c.drawString(40, y, f'Sala: {room.code}')
    y -= 20
    c.setFont('Helvetica-Bold', 11)
    c.drawString(40, y, 'Presença consolidada')
    y -= 18
    sub = db.session.query(
        AttendanceLog.participant_id,
        func.min(AttendanceLog.entered_at).label('first_in'),
        func.max(AttendanceLog.exited_at).label('last_out')
    ).filter_by(room_id=room.id).group_by(AttendanceLog.participant_id).all()
    c.setFont('Helvetica', 9)
    for item in sub:
        p = Participant.query.get(item.participant_id)
        if not p:
            continue
        c.drawString(40, y, f'{p.full_name} | entrada: {item.first_in or "-"} | saída: {item.last_out or "-"}'[:110])
        y -= 14
        if y < 90:
            c.showPage()
            y = 800
            c.setFont('Helvetica', 9)

    c.showPage()
    y = 800
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, 'Resumo da reunião')
    y -= 20
    c.setFont('Helvetica', 10)
    for line in (room.summary_text or '').splitlines() or ['']:
        c.drawString(40, y, line[:100])
        y -= 14
    y -= 10
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y, 'Decisões')
    y -= 20
    c.setFont('Helvetica', 10)
    for line in (room.decisions_text or '').splitlines() or ['']:
        c.drawString(40, y, line[:100])
        y -= 14

    vote = VoteSession.query.filter_by(room_id=room.id).order_by(VoteSession.id.desc()).first()
    if vote:
        tally = tally_vote(vote, room)
        c.showPage()
        y = 800
        c.setFont('Helvetica-Bold', 12)
        c.drawString(40, y, f'Votação: {vote.title}')
        y -= 20
        c.setFont('Helvetica', 10)
        c.drawString(40, y, f'Resultado: {tally["result"]}')
        y -= 16
        c.drawString(40, y, f'Presentes: {tally["presentes"]} | Aptos: {tally["aptos"]} | Votaram: {tally["votaram"]} | Faltam: {tally["faltam"]}')
        y -= 20
        for opt, count in tally['counts'].items():
            c.drawString(40, y, f'{opt}: {count} ({tally["percentages"].get(opt, 0)}%)')
            y -= 14
        if tally['details']:
            y -= 8
            for detail in tally['details']:
                c.drawString(40, y, f'{detail["name"]}: {detail["option"]}'[:110])
                y -= 14
                if y < 70:
                    c.showPage()
                    y = 800
                    c.setFont('Helvetica', 10)

    c.save()
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'{room.code}.pdf')


@socketio.on('join_room')
def on_join(data):
    token = (data or {}).get('join_token')
    p = Participant.query.filter_by(join_token=token).first()
    if not p or p.removed:
        emit('removed', {'reason': 'Acesso inválido para esta sala.'})
        disconnect()
        return
    room = MeetingRoom.query.get_or_404(p.room_id)
    sid_to_participant[request.sid] = p.id
    join_room(room_channel(room.code))
    join_room(participant_channel(p))
    p.online = True
    p.left_at = None
    db.session.commit()
    db.session.add(AttendanceLog(room_id=room.id, participant_id=p.id))
    db.session.commit()
    emit('joined_ok', {'participant_id': p.id, 'room_code': room.code})
    emit_room_state(room.code)


@socketio.on('disconnect')
def on_disconnect():
    pid = sid_to_participant.pop(request.sid, None)
    if not pid:
        return
    p = Participant.query.get(pid)
    if not p:
        return
    p.online = False
    p.left_at = utcnow()
    db.session.commit()
    log = AttendanceLog.query.filter_by(participant_id=pid, exited_at=None).order_by(AttendanceLog.id.desc()).first()
    if log:
        log.exited_at = utcnow()
        db.session.commit()
    room = MeetingRoom.query.get(p.room_id)
    if room:
        emit_room_state(room.code)


@socketio.on('raise_hand')
def on_raise_hand(data):
    p = Participant.query.filter_by(join_token=(data or {}).get('join_token')).first()
    if not p or p.removed:
        return
    room = MeetingRoom.query.get_or_404(p.room_id)
    runtime = room_runtime[room.code]
    if p.id not in runtime['hands']:
        runtime['hands'].append(p.id)
    emit_room_state(room.code)


@socketio.on('lower_hand')
def on_lower_hand(data):
    p = Participant.query.filter_by(join_token=(data or {}).get('join_token')).first()
    if not p or p.removed:
        return
    room = MeetingRoom.query.get_or_404(p.room_id)
    runtime = room_runtime[room.code]
    if p.id in runtime['hands']:
        runtime['hands'].remove(p.id)
    emit_room_state(room.code)


@socketio.on('speaker_update')
def on_speaker_update(data):
    p = Participant.query.filter_by(join_token=(data or {}).get('join_token')).first()
    if not p or p.removed:
        return
    room = MeetingRoom.query.get_or_404(p.room_id)
    runtime = room_runtime[room.code]
    runtime['speaker_id'] = p.id if bool((data or {}).get('speaking')) else (None if runtime['speaker_id'] == p.id else runtime['speaker_id'])
    emit_room_state(room.code)


@socketio.on('screen_share')
def on_screen_share(data):
    p = Participant.query.filter_by(join_token=(data or {}).get('join_token')).first()
    if not p or p.removed:
        return
    room = MeetingRoom.query.get_or_404(p.room_id)
    runtime = room_runtime[room.code]
    runtime['screen_share_id'] = p.id if bool((data or {}).get('active')) else (None if runtime.get('screen_share_id') == p.id else runtime.get('screen_share_id'))
    runtime['selected_id'] = runtime['screen_share_id'] or runtime.get('selected_id')
    emit_room_state(room.code)


@socketio.on('signal')
def on_signal(data):
    sender = Participant.query.filter_by(join_token=(data or {}).get('join_token')).first()
    target_id = (data or {}).get('target_id')
    target = Participant.query.get(target_id) if target_id else None
    if not sender or sender.removed or not target or target.removed:
        return
    emit('signal', {
        'from_id': sender.id,
        'type': data.get('type'),
        'description': data.get('description'),
        'candidate': data.get('candidate'),
        'sender_name': sender.display_name,
    }, to=participant_channel(target))


with app.app_context():
    os.makedirs(os.path.join(BASE_DIR, 'instance'), exist_ok=True)
    db.create_all()
    ensure_schema()


if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=int(os.getenv('PORT', 5000)), debug=True)
